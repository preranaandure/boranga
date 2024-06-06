import json
import logging
from io import BytesIO

import pandas as pd
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from ledger_api_client.ledger_models import EmailUserRO as EmailUser
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from rest_framework import mixins, serializers, views, viewsets
from rest_framework.decorators import action as detail_route
from rest_framework.decorators import action as list_route
from rest_framework.decorators import renderer_classes
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework_datatables.filters import DatatablesFilterBackend
from rest_framework_datatables.pagination import DatatablesPageNumberPagination

from boranga import exceptions
from boranga.components.conservation_status.models import (
    CommonwealthConservationList,
    ConservationChangeCode,
    ConservationStatus,
    ConservationStatusAmendmentRequest,
    ConservationStatusAmendmentRequestDocument,
    ConservationStatusDocument,
    ConservationStatusReferral,
    ConservationStatusUserAction,
    IUCNVersion,
    ProposalAmendmentReason,
    WALegislativeCategory,
    WALegislativeList,
    WAPriorityCategory,
    WAPriorityList,
)
from boranga.components.conservation_status.serializers import (
    ConservationStatusAmendmentRequestDisplaySerializer,
    ConservationStatusAmendmentRequestSerializer,
    ConservationStatusDocumentSerializer,
    ConservationStatusLogEntrySerializer,
    ConservationStatusReferralSerializer,
    ConservationStatusSerializer,
    ConservationStatusUserActionSerializer,
    CreateConservationStatusSerializer,
    DTConservationStatusReferralSerializer,
    InternalConservationStatusSerializer,
    InternalSaveConservationStatusDocumentSerializer,
    ListCommunityConservationStatusSerializer,
    ListConservationStatusSerializer,
    ListSpeciesConservationStatusSerializer,
    ProposedApprovalSerializer,
    ProposedDeclineSerializer,
    SaveCommunityConservationStatusSerializer,
    SaveConservationStatusDocumentSerializer,
    SaveSpeciesConservationStatusSerializer,
    SendReferralSerializer,
)
from boranga.components.conservation_status.utils import cs_proposal_submit
from boranga.components.main.related_item import RelatedItemsSerializer
from boranga.components.species_and_communities.models import (
    ClassificationSystem,
    CommunityTaxonomy,
    GroupType,
    Species,
    Taxonomy,
    TaxonVernacular,
)
from boranga.components.users.models import SubmitterCategory
from boranga.helpers import is_external_contributor, is_internal

logger = logging.getLogger(__name__)


# used for external CS Dash filter
class GetConservationListDict(views.APIView):
    def get(self, request, format=None):
        # TODO: Check where this is used and add group type filtering if necessary
        group_type = None
        res_json = {
            "wa_priority_lists": WAPriorityList.get_lists_dict(group_type),
            "wa_priority_categories": WAPriorityCategory.get_categories_dict(
                group_type
            ),
            "wa_legislative_lists": WALegislativeList.get_lists_dict(group_type),
            "wa_legislative_categories": WALegislativeCategory.get_categories_dict(
                group_type
            ),
            "commonwealth_conservation_lists": CommonwealthConservationList.get_lists_dict(
                group_type
            ),
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")


class GetSpeciesDisplay(views.APIView):
    def get(self, request, format=None):
        # requires species_id or taxon_id
        # TODO: remove species id once scientific_name_lookup changes applied to both CS and OCR/OCC
        # no auth should be required here
        res_json = {}

        species_id = request.GET.get("species_id", "")
        taxon_id = request.GET.get("taxon_id", "")

        if species_id and not taxon_id:
            species = Species.objects.filter(id=species_id)
            if species.exists() and species.first().taxonomy:
                res_json["species_id"] = species.first().id
                taxon_id = species.first().taxonomy.id

        if taxon_id:
            taxonomy = Taxonomy.objects.filter(id=taxon_id)
            if taxonomy.exists():
                res_json["id"] = taxon_id
                res_json["name"] = taxonomy.first().scientific_name
                res_json["taxon_previous_name"] = taxonomy.first().taxon_previous_name
                res_json["common_name"] = taxonomy.first().taxon_vernacular_name

        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")


class GetCommunityDisplay(views.APIView):
    def get(self, request, format=None):
        # requires community_id
        # no auth should be required here
        res_json = {}

        community_id = request.GET.get("community_id", "")

        if community_id:
            community_taxon = CommunityTaxonomy.objects.filter(
                community_id=community_id
            )
            if community_taxon.exists():
                res_json["id"] = community_id
                res_json["name"] = community_taxon.first().community_name

        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")


class GetCSProfileDict(views.APIView):
    def get(self, request, format=None):
        group_type = request.GET.get("group_type", "")

        # NOTE: getting all (non-draft) species and communities here may not be a good idea
        # - moving these to their own endpoint (with an id req)
        # species_list = []
        # if group_type:
        #    exculde_status = ["draft"]
        #    species = Species.objects.filter(
        #        ~Q(processing_status__in=exculde_status)
        #        & ~Q(taxonomy=None)
        #        & Q(group_type__name=group_type)
        #    )
        #    if species:
        #        for specimen in species:
        #            species_list.append(
        #                {
        #                    "id": specimen.id,
        #                    "name": specimen.taxonomy.scientific_name,
        #                    "taxon_previous_name": specimen.taxonomy.taxon_previous_name,
        #                }
        #            )
        # community_list = []
        # exculde_status = ["draft"]
        # communities = CommunityTaxonomy.objects.filter(
        #    ~Q(community__processing_status__in=exculde_status)
        # )  # TODO remove later as every community will have community name
        # if communities:
        #    for specimen in communities:
        #        community_list.append(
        #            {
        #                "id": specimen.community.id,
        #                "name": specimen.community_name,
        #            }
        #        )

        iucn_version_list = []
        if group_type:
            versions = IUCNVersion.objects.filter()
            if versions:
                for option in versions:
                    iucn_version_list.append(
                        {
                            "id": option.id,
                            "code": option.code,
                        }
                    )
        res_json = {
            # "species_list": species_list,
            # "community_list": community_list,
            "wa_priority_lists": WAPriorityList.get_lists_dict(group_type),
            "wa_priority_categories": WAPriorityCategory.get_categories_dict(
                group_type
            ),
            "wa_legislative_lists": WALegislativeList.get_lists_dict(group_type),
            "wa_legislative_categories": WALegislativeCategory.get_categories_dict(
                group_type
            ),
            "commonwealth_conservation_lists": CommonwealthConservationList.get_lists_dict(
                group_type
            ),
            "iucn_version_list": iucn_version_list,
            "change_code_list": ConservationChangeCode.get_filter_list(),
            "submitter_categories": SubmitterCategory.get_filter_list(),
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")


class SpeciesConservationStatusFilterBackend(DatatablesFilterBackend):
    # TODO - Instead of having all the if else statements just make two separate classes
    # and then do the splitting logic in the get_filter_class of the viewset...
    def filter_queryset(self, request, queryset, view):
        total_count = queryset.count()

        # filter_group_type
        filter_group_type = request.POST.get("filter_group_type")
        if queryset.model is ConservationStatus:
            if filter_group_type:
                # queryset = queryset.filter(species__group_type__name=filter_group_type)
                # changed to application_type (ie group_type)
                queryset = queryset.filter(application_type__name=filter_group_type)
        elif queryset.model is ConservationStatusReferral:
            if filter_group_type:
                # queryset = queryset.filter(species__group_type__name=filter_group_type)
                # changed to application_type (ie group_type)
                queryset = queryset.filter(
                    conservation_status__application_type__name=filter_group_type
                )

        # filter_scientific_name
        filter_scientific_name = request.POST.get("filter_scientific_name")
        if queryset.model is ConservationStatus:
            if filter_scientific_name and not filter_scientific_name.lower() == "all":
                queryset = queryset.filter(species_taxonomy=filter_scientific_name)
        elif queryset.model is ConservationStatusReferral:
            if filter_scientific_name and not filter_scientific_name.lower() == "all":
                queryset = queryset.filter(
                    conservation_status__species_taxonomy=filter_scientific_name
                )

        filter_common_name = request.POST.get("filter_common_name")
        if queryset.model is ConservationStatus:
            if filter_common_name and not filter_common_name.lower() == "all":
                queryset = queryset.filter(
                    species__taxonomy__vernaculars__id=filter_common_name
                )
        elif queryset.model is ConservationStatusReferral:
            if filter_common_name and not filter_common_name.lower() == "all":
                queryset = queryset.filter(
                    conservation_status__species__taxonomy__vernaculars__id=filter_common_name
                )

        filter_phylogenetic_group = request.POST.get("filter_phylogenetic_group")
        if queryset.model is ConservationStatus:
            if (
                filter_phylogenetic_group
                and not filter_phylogenetic_group.lower() == "all"
            ):
                queryset = queryset.filter(
                    species__taxonomy__informal_groups__classification_system_fk_id=filter_phylogenetic_group
                )
        elif queryset.model is ConservationStatusReferral:
            if (
                filter_phylogenetic_group
                and not filter_phylogenetic_group.lower() == "all"
            ):
                queryset = queryset.filter(
                    conservation_status__species__taxonomy__informal_groups__classification_system_id=filter_phylogenetic_group  # noqa
                )

        filter_family = request.POST.get("filter_family")
        if queryset.model is ConservationStatus:
            if filter_family and not filter_family.lower() == "all":
                queryset = queryset.filter(species__taxonomy__family_id=filter_family)
        elif queryset.model is ConservationStatusReferral:
            if filter_family and not filter_family.lower() == "all":
                queryset = queryset.filter(
                    conservation_status__species__taxonomy__family_id=filter_family
                )

        filter_genus = request.POST.get("filter_genus")
        if queryset.model is ConservationStatus:
            if filter_genus and not filter_genus.lower() == "all":
                queryset = queryset.filter(species__taxonomy__genera_id=filter_genus)
        elif queryset.model is ConservationStatusReferral:
            if filter_genus and not filter_genus.lower() == "all":
                queryset = queryset.filter(
                    conservation_status__species__taxonomy__genera_id=filter_genus
                )

        filter_change_code = request.POST.get("filter_change_code")
        if queryset.model is ConservationStatus:
            if filter_change_code and not filter_change_code.lower() == "all":
                queryset = queryset.filter(change_code__id=filter_change_code)
        elif queryset.model is ConservationStatusReferral:
            if filter_change_code and not filter_change_code.lower() == "all":
                queryset = queryset.filter(
                    conservation_status__change_code__id=filter_change_code
                )

        filter_wa_legislative_list = request.POST.get("filter_wa_legislative_list")
        if (
            filter_wa_legislative_list
            and not filter_wa_legislative_list.lower() == "all"
        ):
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(
                    wa_legislative_list=filter_wa_legislative_list
                ).distinct()
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__wa_legislative_list=filter_wa_legislative_list
                )

        filter_wa_legislative_category = request.POST.get(
            "filter_wa_legislative_category"
        )
        if (
            filter_wa_legislative_category
            and not filter_wa_legislative_category.lower() == "all"
        ):
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(
                    wa_legislative_category=filter_wa_legislative_category
                ).distinct()
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__wa_legislative_category=filter_wa_legislative_category
                )

        filter_wa_priority_category = request.POST.get("filter_wa_priority_category")
        if (
            filter_wa_priority_category
            and not filter_wa_priority_category.lower() == "all"
        ):
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(
                    wa_priority_category=filter_wa_priority_category
                ).distinct()
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__wa_priority_category=filter_wa_priority_category
                )

        filter_commonwealth_relevance = request.POST.get(
            "filter_commonwealth_relevance"
        )
        if filter_commonwealth_relevance == "true":
            if queryset.model is ConservationStatus:
                queryset = queryset.exclude(commonwealth_conservation_list__isnull=True)
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.exclude(
                    conservation_status__commonwealth_conservation_list__isnull=True
                )

        filter_international_relevance = request.POST.get(
            "filter_international_relevance"
        )
        if filter_international_relevance == "true":
            if queryset.model is ConservationStatus:
                queryset = queryset.exclude(international_conservation__isnull=True)
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.exclude(
                    conservation_status__international_conservation__isnull=True
                )

        filter_from_effective_from_date = request.POST.get(
            "filter_from_effective_from_date"
        )
        filter_to_effective_from_date = request.POST.get(
            "filter_to_effective_from_date"
        )
        filter_from_review_due_date = request.POST.get("filter_from_review_due_date")

        filter_from_effective_to_date = request.POST.get(
            "filter_from_effective_to_date"
        )
        filter_to_effective_to_date = request.POST.get("filter_to_effective_to_date")
        filter_to_review_due_date = request.POST.get("filter_to_review_due_date")

        if queryset.model is ConservationStatus:
            if filter_from_effective_from_date:
                queryset = queryset.filter(
                    conservationstatusissuanceapprovaldetails__effective_from_date__gte=filter_from_effective_from_date
                )
            if filter_to_effective_from_date:
                queryset = queryset.filter(
                    conservationstatusissuanceapprovaldetails__effective_from_date__lte=filter_to_effective_from_date
                )
            if filter_from_effective_to_date:
                queryset = queryset.filter(
                    conservationstatusissuanceapprovaldetails__effective_to_date__gte=filter_from_effective_to_date
                )
            if filter_to_effective_to_date:
                queryset = queryset.filter(
                    conservationstatusissuanceapprovaldetails__effective_to_date__lte=filter_to_effective_to_date
                )
            if filter_from_review_due_date:
                queryset = queryset.filter(
                    review_due_date__gte=filter_from_review_due_date
                )
            if filter_to_review_due_date:
                queryset = queryset.filter(
                    review_due_date__lte=filter_to_review_due_date
                )

        elif queryset.model is ConservationStatusReferral:

            if filter_from_effective_from_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__effective_from_date__gte=filter_from_effective_from_date  # noqa
                )
            if filter_to_effective_from_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__effective_from_date__lte=filter_to_effective_from_date  # noqa
                )
            if filter_from_effective_to_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__effective_from_date__gte=filter_from_effective_to_date  # noqa
                )
            if filter_to_effective_to_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__effective_from_date__lte=filter_to_effective_to_date  # noqa
                )
            if filter_from_review_due_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__review_due_date__gte=filter_from_review_due_date  # noqa
                )
            if filter_to_review_due_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__review_due_date__lte=filter_to_review_due_date  # noqa
                )

        filter_application_status = request.POST.get("filter_application_status")
        if queryset.model is ConservationStatus:
            if (
                filter_application_status
                and not filter_application_status.lower() == "all"
            ):
                queryset = queryset.filter(processing_status=filter_application_status)
        elif queryset.model is ConservationStatusReferral:
            if (
                filter_application_status
                and not filter_application_status.lower() == "all"
            ):
                queryset = queryset.filter(
                    conservation_status__processing_status=filter_application_status
                )

        filter_assessor = request.POST.get("filter_assessor")
        if filter_assessor and not filter_assessor.lower() == "all":
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(assigned_officer=filter_assessor)
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__assigned_officer=filter_assessor
                )

        filter_submitter = request.POST.get("filter_submitter")
        if filter_submitter and not filter_submitter.lower() == "all":
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(submitter=filter_submitter)
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__submitter=filter_submitter
                )

        filter_submitter_category = request.POST.get("filter_submitter_category")
        if queryset.model is ConservationStatus:
            if (
                filter_submitter_category
                and not filter_submitter_category.lower() == "all"
            ):
                queryset = queryset.filter(
                    submitter_information__submitter_category__id=filter_submitter_category
                )
        elif queryset.model is ConservationStatusReferral:
            if (
                filter_submitter_category
                and not filter_submitter_category.lower() == "all"
            ):
                queryset = queryset.filter(
                    conservation_status__submitter_information__submitter_category__id=filter_submitter_category
                )

        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        try:
            queryset = super().filter_queryset(request, queryset, view)
        except Exception as e:
            print(e)
        setattr(view, "_datatables_total_count", total_count)
        return queryset


class SpeciesConservationStatusPaginatedViewSet(viewsets.ReadOnlyModelViewSet):
    filter_backends = (SpeciesConservationStatusFilterBackend,)
    pagination_class = DatatablesPageNumberPagination
    queryset = ConservationStatus.objects.none()
    serializer_class = ListSpeciesConservationStatusSerializer
    page_size = 10

    def get_queryset(self):
        # request_user = self.request.user
        qs = ConservationStatus.objects.none()

        if is_internal(self.request):
            qs = ConservationStatus.objects.exclude(
                processing_status=ConservationStatus.PROCESSING_STATUS_DISCARDED
            )

        return qs

    @list_route(
        methods=[
            "GET",  # still allow get for local development speed
            "POST",
        ],
        detail=False,
    )
    def species_cs_internal(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListSpeciesConservationStatusSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    # used for Meeting Agenda Modal where status is 'ready_for_agenda'
    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def agenda_cs_internal(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = qs.filter(
            processing_status=ConservationStatus.PROCESSING_STATUS_READY_FOR_AGENDA
        )
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListSpeciesConservationStatusSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def species_cs_internal_export(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.POST.get("export_format")
        allowed_fields = [
            "species_number",
            "scientific_name",
            "common_name",
            "family",
            "genus",
            "phylogenetic_group",
            "conservation_list",
            "conservation_category",
            "processing_status",
            "effective_from_date",
            "effective_to_date",
            "conservation_status_number",
        ]

        serializer = ListSpeciesConservationStatusSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Species",
            "Scientific Name",
            "Common Name",
            "Family",
            "Genera",
            "Phylo Group",
            "Processing Status",
            "Effective From Date",
            "Effective To Date",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Species",
            "Scientific Name",
            "Common Name",
            "Effective From Date",
            "Effective To Date",
            "Family",
            "Genera",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ConservationStatus_Species.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ConservationStatus_Species.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def species_cs_referrals_internal(self, request, *args, **kwargs):
        """
        Used by the internal Referred To Me (Flora/Fauna)dashboard

        http://localhost:8499/api/species_conservation_cstatus_paginated/species_cs_referrals_internal/?format=datatables&draw=1&length=2
        """
        self.serializer_class = DTConservationStatusReferralSerializer
        qs = (
            ConservationStatusReferral.objects.filter(referral=request.user.id)
            if is_internal(self.request)
            else ConservationStatusReferral.objects.none()
        )
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = DTConservationStatusReferralSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def species_cs_referrals_internal_export(self, request, *args, **kwargs):

        self.serializer_class = ConservationStatusReferralSerializer
        qs = (
            ConservationStatusReferral.objects.filter(referral=request.user.id)
            if is_internal(self.request)
            else ConservationStatusReferral.objects.none()
        )
        qs = self.filter_queryset(qs)
        export_format = request.POST.get("export_format")
        allowed_fields = [
            "species_number",
            "scientific_name",
            "common_name",
            "family",
            "genus",
            "conservation_list",
            "conservation_category",
            "processing_status",
            "conservation_status_number",
        ]

        serializer = DTConservationStatusReferralSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Processing Status",
            "Number",
            "Species",
            "Scientific Name",
            "Common Name",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Species",
            "Scientific Name",
            "Common Name",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ConservationStatus_Species_Referrals.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ConservationStatus_Species_Referrals.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")


class CommunityConservationStatusFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):
        total_count = queryset.count()

        # filter_group_type
        filter_group_type = request.POST.get("filter_group_type")
        if filter_group_type:
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(application_type__name=filter_group_type)
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__application_type__name=filter_group_type
                )

        # filter_community_migrated_id
        filter_community_migrated_id = request.POST.get("filter_community_migrated_id")
        if (
            filter_community_migrated_id
            and not filter_community_migrated_id.lower() == "all"
        ):
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(
                    community__taxonomy__id=filter_community_migrated_id
                )
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__community__taxonomy__id=filter_community_migrated_id
                )

        # filter_community_name
        filter_community_name = request.POST.get("filter_community_name")
        if filter_community_name and not filter_community_name.lower() == "all":
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(
                    community__taxonomy__id=filter_community_name
                )
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__community__taxonomy__id=filter_community_name
                )

        filter_conservation_list = request.POST.get("filter_conservation_list")
        if filter_conservation_list and not filter_conservation_list.lower() == "all":
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(conservation_list=filter_conservation_list)
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__conservation_list=filter_conservation_list
                )

        filter_conservation_category = request.POST.get("filter_conservation_category")
        if (
            filter_conservation_category
            and not filter_conservation_category.lower() == "all"
        ):
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(
                    conservation_category=filter_conservation_category
                )
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__conservation_category=filter_conservation_category
                )

        filter_region = request.POST.get("filter_region")
        if filter_region and not filter_region.lower() == "all":
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(community__region=filter_region)
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__community__region=filter_region
                )

        filter_district = request.POST.get("filter_district")
        if filter_district and not filter_district.lower() == "all":
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(community__district=filter_district)
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__community__district=filter_district
                )

        filter_from_effective_from_date = request.POST.get(
            "filter_from_effective_from_date"
        )
        filter_to_effective_from_date = request.POST.get(
            "filter_to_effective_from_date"
        )

        filter_from_effective_to_date = request.POST.get(
            "filter_from_effective_to_date"
        )
        filter_to_effective_to_date = request.POST.get("filter_to_effective_to_date")
        if queryset.model is ConservationStatus:
            if filter_from_effective_from_date:
                queryset = queryset.filter(
                    conservationstatusissuanceapprovaldetails__effective_from_date__gte=filter_from_effective_from_date
                )
            if filter_to_effective_from_date:
                queryset = queryset.filter(
                    conservationstatusissuanceapprovaldetails__effective_from_date__lte=filter_to_effective_from_date
                )

            if filter_from_effective_to_date:
                queryset = queryset.filter(
                    conservationstatusissuanceapprovaldetails__effective_to_date__gte=filter_from_effective_to_date
                )
            if filter_to_effective_to_date:
                queryset = queryset.filter(
                    conservationstatusissuanceapprovaldetails__effective_to_date__lte=filter_to_effective_to_date
                )

        elif queryset.model is ConservationStatusReferral:

            if filter_from_effective_from_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__effective_from_date__gte=filter_from_effective_from_date  # noqa
                )
            if filter_to_effective_from_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__effective_from_date__lte=filter_to_effective_from_date  # noqa
                )

            if filter_from_effective_to_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__effective_from_date__gte=filter_from_effective_to_date  # noqa
                )
            if filter_to_effective_to_date:
                queryset = queryset.filter(
                    conservation_status__conservationstatusissuanceapprovaldetails__effective_from_date__lte=filter_to_effective_to_date  # noqa
                )

        filter_application_status = request.POST.get("filter_application_status")
        if filter_application_status and not filter_application_status.lower() == "all":
            if queryset.model is ConservationStatus:
                queryset = queryset.filter(processing_status=filter_application_status)
            elif queryset.model is ConservationStatusReferral:
                queryset = queryset.filter(
                    conservation_status__processing_status=filter_application_status
                )

        # getter = request.query_params.get
        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        try:
            queryset = super().filter_queryset(request, queryset, view)
        except Exception as e:
            print(e)
        setattr(view, "_datatables_total_count", total_count)
        return queryset


class CommunityConservationStatusPaginatedViewSet(viewsets.ReadOnlyModelViewSet):
    filter_backends = (CommunityConservationStatusFilterBackend,)
    pagination_class = DatatablesPageNumberPagination
    # renderer_classes = (CommunityConservationStatusRenderer,)
    queryset = ConservationStatus.objects.none()
    serializer_class = ListCommunityConservationStatusSerializer
    page_size = 10

    def get_queryset(self):
        # request_user = self.request.user
        qs = ConservationStatus.objects.none()

        if is_internal(self.request):
            qs = ConservationStatus.objects.exclude(
                processing_status=ConservationStatus.PROCESSING_STATUS_DISCARDED
            )
        return qs

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def community_cs_internal(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListCommunityConservationStatusSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def agenda_cs_internal(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = qs.filter(
            processing_status=ConservationStatus.PROCESSING_STATUS_READY_FOR_AGENDA
        )
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListCommunityConservationStatusSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def community_cs_internal_export(self, request, *args, **kwargs):

        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.POST.get("export_format")
        allowed_fields = [
            "conservation_status_number",
            "community_number",
            "community_migrated_id",
            "community_name",
            "region",
            "district",
            "conservation_list",
            "conservation_category",
            "processing_status",
            "effective_from_date",
            "effective_to_date",
        ]

        serializer = ListCommunityConservationStatusSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Community",
            "Community Id",
            "Community Name",
            "Region",
            "District",
            "Processing Status",
            "Effective From Date",
            "Effective To Date",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Community",
            "Community Id",
            "Community Name",
            "Region",
            "District",
            "Effective From Date",
            "Effective To Date",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ConservationStatus_Communities.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ConservationStatus_Communities.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def community_cs_referrals_internal(self, request, *args, **kwargs):
        """
        Used by the internal Referred to me dashboard

        http://localhost:8499/api/community_conservation_status_paginated/community_cs_referrals_internal/?format=datatables&draw=1&length=2
        """
        self.serializer_class = ConservationStatusReferralSerializer
        qs = (
            ConservationStatusReferral.objects.filter(referral=request.user.id)
            if is_internal(self.request)
            else ConservationStatusReferral.objects.none()
        )

        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = DTConservationStatusReferralSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def community_cs_referrals_internal_export(self, request, *args, **kwargs):

        self.serializer_class = ConservationStatusReferralSerializer
        qs = (
            ConservationStatusReferral.objects.filter(referral=request.user.id)
            if is_internal(self.request)
            else ConservationStatusReferral.objects.none()
        )
        qs = self.filter_queryset(qs)
        export_format = request.POST.get("export_format")
        allowed_fields = [
            "conservation_list",
            "conservation_category",
            "processing_status",
            "community_number",
            "community_migrated_id",
            "community_name",
            "conservation_status_number",
        ]

        serializer = DTConservationStatusReferralSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Processing Status",
            "Number",
            "Community",
            "Community Id",
            "Community Name",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Community",
            "Community Id",
            "Community Name",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ConservationStatus_Communities_Referrals.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ConservationStatus_Communities_Referrals.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")


class ConservationStatusFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):
        total_count = queryset.count()

        filter_group_type = request.GET.get("filter_group_type")
        # if filter_group_type and filter_group_type == flora or filter_group_type == fauna:
        #     queryset = queryset.filter(species__group_type__name=filter_group_type)
        # elif filter_group_type and filter_group_type == community:
        #     queryset = queryset.filter(community__group_type__name=filter_group_type)
        if filter_group_type and not filter_group_type.lower() == "all":
            queryset = queryset.filter(application_type__name=filter_group_type)

        # filter_scientific_name is the species_id
        filter_scientific_name = request.GET.get("filter_scientific_name")
        if filter_scientific_name and not filter_scientific_name.lower() == "all":
            queryset = queryset.filter(species_taxonomy=filter_scientific_name)

        # filter_community_name is the community_id
        filter_community_name = request.GET.get("filter_community_name")
        if filter_community_name and not filter_community_name.lower() == "all":
            queryset = queryset.filter(community=filter_community_name)

        filter_conservation_list = request.GET.get("filter_conservation_list")
        if filter_conservation_list and not filter_conservation_list.lower() == "all":
            queryset = queryset.filter(conservation_list=filter_conservation_list)

        filter_conservation_category = request.GET.get("filter_conservation_category")
        if (
            filter_conservation_category
            and not filter_conservation_category.lower() == "all"
        ):
            queryset = queryset.filter(
                conservation_category=filter_conservation_category
            )

        filter_application_status = request.GET.get("filter_application_status")
        if filter_application_status and not filter_application_status.lower() == "all":
            queryset = queryset.filter(customer_status=filter_application_status)

        # getter = request.query_params.get
        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        try:
            queryset = super().filter_queryset(request, queryset, view)
        except Exception as e:
            print(e)
        setattr(view, "_datatables_total_count", total_count)
        return queryset


class ConservationStatusPaginatedViewSet(viewsets.ReadOnlyModelViewSet):
    filter_backends = (ConservationStatusFilterBackend,)
    pagination_class = DatatablesPageNumberPagination
    queryset = ConservationStatus.objects.none()
    serializer_class = ListConservationStatusSerializer
    page_size = 10

    def get_queryset(self):
        qs = self.queryset

        if is_external_contributor(self.request):
            qs = ConservationStatus.objects.filter(submitter=self.request.user.id)

        if is_internal(self.request):
            qs = ConservationStatus.objects.exclude(
                processing_status=ConservationStatus.PROCESSING_STATUS_DISCARDED
            )
        return qs

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def conservation_status_external(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = qs.filter(internal_application=False)
        # TODO Not Sure but to filter for only WA listed conservation lists for external
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListConservationStatusSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)


class ConservationStatusViewSet(viewsets.GenericViewSet, mixins.RetrieveModelMixin):
    queryset = ConservationStatus.objects.none()
    serializer_class = ConservationStatusSerializer
    lookup_field = "id"

    def get_queryset(self):
        qs = self.queryset

        if is_external_contributor(self.request):
            qs = ConservationStatus.objects.filter(submitter=self.request.user.id)

        if is_internal(self.request):
            qs = ConservationStatus.objects.exclude(
                processing_status=ConservationStatus.PROCESSING_STATUS_DISCARDED
            )
        return qs

    def internal_serializer_class(self):
        return InternalConservationStatusSerializer

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def internal_conservation_status(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})

        res_json = {"conservation_status_obj": serializer.data}
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def filter_list(self, request, *args, **kwargs):
        """Used by the Related Items dashboard filters"""
        related_type = ConservationStatus.RELATED_ITEM_CHOICES
        res_json = json.dumps(related_type)
        return HttpResponse(res_json, content_type="application/json")

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def conservation_status_save(self, request, *args, **kwargs):
        instance = self.get_object()
        request_data = request.data
        # to resolve error for serializer submitter id as object is received in request
        if request_data["submitter"]:
            request.data["submitter"] = "{}".format(request_data["submitter"].get("id"))
        if (
            instance.application_type.name == GroupType.GROUP_TYPE_FLORA
            or instance.application_type.name == GroupType.GROUP_TYPE_FAUNA
        ):
            serializer = SaveSpeciesConservationStatusSerializer(
                instance, data=request_data, partial=True
            )
        elif instance.application_type.name == GroupType.GROUP_TYPE_COMMUNITY:
            serializer = SaveCommunityConservationStatusSerializer(
                instance, data=request_data, partial=True
            )

        serializer.is_valid(raise_exception=True)
        if serializer.is_valid():
            serializer.save(version_user=request.user)

            instance.log_user_action(
                ConservationStatusUserAction.ACTION_SAVE_APPLICATION.format(
                    instance.conservation_status_number
                ),
                request,
            )

        return redirect(reverse("internal"))

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def conservation_status_edit(self, request, *args, **kwargs):
        instance = self.get_object()
        request_data = request.data
        if not instance.can_assess(request):
            raise exceptions.ProposalNotAuthorized()
        if instance.processing_status == "approved":
            # to resolve error for serializer submitter id as object is received in request
            if request_data["submitter"]:
                request.data["submitter"] = "{}".format(
                    request_data["submitter"].get("id")
                )
            if (
                instance.application_type.name == GroupType.GROUP_TYPE_FLORA
                or instance.application_type.name == GroupType.GROUP_TYPE_FAUNA
            ):
                serializer = SaveSpeciesConservationStatusSerializer(
                    instance, data=request_data, partial=True
                )
            elif instance.application_type.name == GroupType.GROUP_TYPE_COMMUNITY:
                serializer = SaveCommunityConservationStatusSerializer(
                    instance, data=request_data, partial=True
                )

            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save(version_user=request.user)

                instance.log_user_action(
                    ConservationStatusUserAction.ACTION_EDIT_APPLICATION.format(
                        instance.conservation_status_number
                    ),
                    request,
                )

        return redirect(reverse("external"))

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def draft(self, request, *args, **kwargs):
        instance = self.get_object()
        request_data = request.data
        if request_data["submitter"]:
            request.data["submitter"] = "{}".format(request_data["submitter"].get("id"))
        if (
            instance.application_type.name == GroupType.GROUP_TYPE_FLORA
            or instance.application_type.name == GroupType.GROUP_TYPE_FAUNA
        ):
            serializer = SaveSpeciesConservationStatusSerializer(
                instance, data=request_data, partial=True
            )
        elif instance.application_type.name == GroupType.GROUP_TYPE_COMMUNITY:
            serializer = SaveCommunityConservationStatusSerializer(
                instance, data=request_data, partial=True
            )

        serializer.is_valid(raise_exception=True)
        if serializer.is_valid():
            serializer.save(version_user=request.user)

        return redirect(reverse("external"))

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    def submit(self, request, *args, **kwargs):
        instance = self.get_object()
        # instance.submit(request,self)
        cs_proposal_submit(instance, request)
        instance.save(version_user=request.user)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        group_type_id = GroupType.objects.get(
            id=request.data.get("application_type_id")
        )
        internal_application = False
        if request.data.get("internal_application"):
            internal_application = request.data.get("internal_application")
        obj = ConservationStatus(
            submitter=request.user.id,
            application_type=group_type_id,
            internal_application=internal_application,
        )
        obj.save(version_user=request.user)
        serialized_obj = CreateConservationStatusSerializer(obj)
        return Response(serialized_obj.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def action_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.action_logs.all()
        serializer = ConservationStatusUserActionSerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def comms_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.comms_logs.all()
        serializer = ConservationStatusLogEntrySerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def add_comms_log(self, request, *args, **kwargs):
        instance = self.get_object()
        mutable = request.data._mutable
        request.data._mutable = True
        request.data["conservation_status"] = f"{instance.id}"
        request.data["staff"] = f"{request.user.id}"
        request.data._mutable = mutable
        serializer = ConservationStatusLogEntrySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        comms = serializer.save()
        # Save the files
        for f in request.FILES:
            document = comms.documents.create()
            document.name = str(request.FILES[f])
            document._file = request.FILES[f]
            document.save()
        # End Save Documents

        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def amendment_request(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.amendment_requests
        qs = qs.filter(status="requested")
        serializer = ConservationStatusAmendmentRequestDisplaySerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def assign_request_user(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.assign_officer(request, request.user)
        # serializer = InternalProposalSerializer(instance,context={'request':request})
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def assign_to(self, request, *args, **kwargs):
        instance = self.get_object()
        user_id = request.data.get("assessor_id", None)
        user = None
        if not user_id:
            raise serializers.ValidationError("An assessor id is required")
        try:
            user = EmailUser.objects.get(id=user_id)
        except EmailUser.DoesNotExist:
            raise serializers.ValidationError(
                "A user with the id passed in does not exist"
            )
        instance.assign_officer(request, user)
        # serializer = InternalProposalSerializer(instance,context={'request':request})
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def unassign(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.unassign(request)
        # serializer = InternalProposalSerializer(instance,context={'request':request})
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    def assesor_send_referral(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = SendReferralSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        # text=serializer.validated_data['text']
        # instance.send_referral(request,serializer.validated_data['email'])
        instance.send_referral(
            request,
            serializer.validated_data["email"],
            serializer.validated_data["text"],
        )
        # serializer = InternalProposalSerializer(instance,context={'request':request})
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    # with new workflow not used at the moment
    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def proposed_decline(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = ProposedDeclineSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.proposed_decline(request, serializer.validated_data)
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def final_decline(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = ProposedDeclineSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.final_decline(request, serializer.validated_data)
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    # with new workflow not used at the moment
    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def proposed_approval(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = ProposedApprovalSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.proposed_approval(request, serializer.validated_data)
        # serializer = InternalProposalSerializer(instance,context={'request':request})
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def final_approval(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = ProposedApprovalSerializer(
            data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        instance.final_approval(request, serializer.validated_data)
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def switch_status(self, request, *args, **kwargs):
        instance = self.get_object()
        status = request.data.get("status")
        approver_comment = request.data.get("approver_comment")

        if not status:
            raise serializers.ValidationError("Status is required")

        if status not in [
            ConservationStatus.PROCESSING_STATUS_WITH_ASSESSOR,
            ConservationStatus.PROCESSING_STATUS_WITH_APPROVER,
        ]:
            raise serializers.ValidationError("The status provided is not allowed")

        instance.move_to_status(request, status, approver_comment)
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def proposed_ready_for_agenda(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.proposed_ready_for_agenda(request)
        # serializer = InternalProposalSerializer(instance,context={'request':request})
        serializer_class = self.internal_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def documents(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.documents.all()
        qs = qs.exclude(
            input_name="conservation_status_approval_doc"
        )  # TODO do we need/not to show approval doc in cs documents tab
        qs = qs.order_by("-uploaded_date")
        serializer = ConservationStatusDocumentSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["patch"], detail=True)
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.discard(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(methods=["patch"], detail=True)
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.reinstate(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(methods=["patch"], detail=True)
    def propose_delist(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.propose_delist(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(methods=["patch"], detail=True)
    def delist(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delist(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_items(self, request, *args, **kwargs):
        instance = self.get_object()
        related_filter_type = request.GET.get("related_filter_type")
        related_items = instance.get_related_items(related_filter_type)
        serializer = RelatedItemsSerializer(related_items, many=True)
        return Response(serializer.data)


class ConservationStatusReferralViewSet(
    viewsets.GenericViewSet, mixins.RetrieveModelMixin
):
    queryset = ConservationStatusReferral.objects.none()
    serializer_class = ConservationStatusReferralSerializer

    def get_queryset(self):
        qs = self.queryset
        if is_internal(self.request):
            qs = ConservationStatusReferral.objects.all()

        return qs

    def get_serializer_class(self):
        return ConservationStatusReferralSerializer

    # used for Species (flora/Fauna)Referred to me internal dashboard filters
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def filter_list(self, request, *args, **kwargs):
        """Used by the internal Referred to me dashboard filters"""
        qs = self.get_queryset()
        group_type = request.GET.get("group_type_name", "")
        species_data_list = []
        if group_type:
            species_qs = (
                qs.filter(conservation_status__species__group_type__name=group_type)
                .values_list("conservation_status__species", flat=True)
                .distinct()
            )
            species = Species.objects.filter(id__in=species_qs)
            if species:
                for i in species:
                    species_data_list.append(
                        {
                            "species_id": i.id,
                        }
                    )
        scientific_name_list = []
        if group_type:
            taxonomy_qs = (
                qs.filter(conservation_status__species__group_type__name=group_type)
                .values_list("conservation_status__species__taxonomy", flat=True)
                .distinct()
            )
            names = Taxonomy.objects.filter(
                id__in=taxonomy_qs
            )  # TODO will need to filter according to  group  selection
            if names:
                for name in names:
                    scientific_name_list.append(
                        {
                            "id": name.id,
                            "name": name.scientific_name,
                        }
                    )
        common_name_list = []
        if group_type:
            taxonomy_qs = (
                qs.filter(conservation_status__species__group_type__name=group_type)
                .values_list("conservation_status__species__taxonomy", flat=True)
                .distinct()
            )
            common_names = TaxonVernacular.objects.filter(taxonomy__in=taxonomy_qs)
            if common_names:
                for name in common_names:
                    common_name_list.append(
                        {
                            "id": name.id,
                            "name": name.vernacular_name,
                        }
                    )
        family_list = []
        if group_type:
            taxonomy_qs = (
                qs.filter(conservation_status__species__group_type__name=group_type)
                .values_list("conservation_status__species__taxonomy", flat=True)
                .distinct()
            )
            families_qs = (
                Taxonomy.objects.filter(Q(id__in=taxonomy_qs))
                .order_by()
                .values_list("family_id", flat=True)
                .distinct()
            )  # fetch all distinct the family_nid(taxon_name_id) for each taxon
            families = Taxonomy.objects.filter(id__in=families_qs)

            if families:
                for family in families:
                    family_list.append(
                        {
                            "id": family.id,
                            "name": family.scientific_name,
                        }
                    )
        phylogenetic_group_list = []
        if group_type:
            taxonomy_qs = (
                qs.filter(conservation_status__species__group_type__name=group_type)
                .values_list("conservation_status__species__taxonomy", flat=True)
                .distinct()
            )
            phylo_group_qs = (
                Taxonomy.objects.filter(
                    Q(id__in=taxonomy_qs) & ~Q(informal_groups=None)
                )
                .values_list("informal_groups__classification_system_fk", flat=True)
                .distinct()
            )
            phylo_groups = ClassificationSystem.objects.filter(
                id__in=phylo_group_qs
            )  # TODO will need to filter according to  group  selection
            if phylo_groups:
                for group in phylo_groups:
                    phylogenetic_group_list.append(
                        {
                            "id": group.id,
                            "name": group.class_desc,
                        }
                    )

        processing_status_list = []
        processing_statuses = (
            qs.filter(conservation_status__processing_status__isnull=False)
            .order_by("conservation_status__processing_status")
            .distinct("conservation_status__processing_status")
            .values_list("conservation_status__processing_status", flat=True)
        )
        if processing_statuses:
            for status in processing_statuses:
                processing_status_list.append(
                    {
                        "value": status,
                        "name": "{}".format(" ".join(status.split("_")).capitalize()),
                    }
                )
        res_json = {
            "species_data_list": species_data_list,
            "scientific_name_list": scientific_name_list,
            "common_name_list": common_name_list,
            "family_list": family_list,
            "phylogenetic_group_list": phylogenetic_group_list,
            "wa_priority_lists": WAPriorityList.get_lists_dict(group_type),
            "wa_priority_categories": WAPriorityCategory.get_categories_dict(
                group_type
            ),
            "wa_legislative_lists": WALegislativeList.get_lists_dict(group_type),
            "wa_legislative_categories": WALegislativeCategory.get_categories_dict(
                group_type
            ),
            "commonwealth_conservation_lists": CommonwealthConservationList.get_lists_dict(
                group_type
            ),
            "processing_status_list": processing_status_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def community_filter_list(self, request, *args, **kwargs):
        """Used by the internal referred to me dashboard filters for community"""
        qs = self.get_queryset()
        group_type = request.GET.get("group_type_name", "")
        community_data_list = []
        if group_type:
            taxonomy_qs = (
                qs.filter(conservation_status__community__isnull=False)
                .values_list("conservation_status__community__taxonomy", flat=True)
                .distinct()
            )
            names = CommunityTaxonomy.objects.filter(
                id__in=taxonomy_qs
            )  # TODO will need to filter according to  group  selection
            if names:
                for name in names:
                    community_data_list.append(
                        {
                            "id": name.id,
                            "community_migrated_id": name.community_migrated_id,
                            "community_name": name.community_name,
                        }
                    )

        processing_status_list = []
        processing_statuses = (
            qs.filter(conservation_status__processing_status__isnull=False)
            .order_by("conservation_status__processing_status")
            .distinct("conservation_status__processing_status")
            .values_list("conservation_status__processing_status", flat=True)
        )
        if processing_statuses:
            for status in processing_statuses:
                processing_status_list.append(
                    {
                        "value": status,
                        "name": "{}".format(" ".join(status.split("_")).capitalize()),
                    }
                )
        res_json = {
            "community_data_list": community_data_list,
            "wa_priority_lists": WAPriorityList.get_lists_dict(group_type),
            "wa_priority_categories": WAPriorityCategory.get_categories_dict(
                group_type
            ),
            "wa_legislative_lists": WALegislativeList.get_lists_dict(group_type),
            "wa_legislative_categories": WALegislativeCategory.get_categories_dict(
                group_type
            ),
            "commonwealth_conservation_lists": CommonwealthConservationList.get_lists_dict(
                group_type
            ),
            "processing_status_list": processing_status_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def datatable_list(self, request, *args, **kwargs):
        conservation_status = request.GET.get("conservation_status", None)
        qs = self.get_queryset().all()
        if conservation_status:
            qs = qs.filter(conservation_status_id=int(conservation_status))
        serializer = DTConservationStatusReferralSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def referral_list(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = self.get_queryset().all()
        qs = qs.filter(
            sent_by=instance.referral, conservation_status=instance.conservation_status
        )
        serializer = DTConservationStatusReferralSerializer(
            qs, many=True, context={"request": request}
        )
        # serializer = ProposalReferralSerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(methods=["GET", "POST"], detail=True)
    def complete(self, request, *args, **kwargs):
        instance = self.get_object()
        # referral_comment = request.data.get('referral_comment')
        # instance.complete(request, referral_comment)
        instance.complete(request)
        serializer = self.get_serializer(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def remind(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.remind(request)
        serializer = InternalConservationStatusSerializer(
            instance.conservation_status, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def recall(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.recall(request)
        serializer = InternalConservationStatusSerializer(
            instance.conservation_status, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def resend(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.resend(request)
        serializer = InternalConservationStatusSerializer(
            instance.conservation_status, context={"request": request}
        )
        return Response(serializer.data)

    # used on referral form
    @detail_route(methods=["post"], detail=True)
    def send_referral(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = SendReferralSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        instance.send_referral(
            request,
            serializer.validated_data["email"],
            serializer.validated_data["text"],
        )
        serializer = self.get_serializer(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def conservation_status_referral_save(self, request, *args, **kwargs):
        instance = self.get_object()
        request_data = request.data
        instance.referral_comment = request_data.get("referral_comment")
        instance.save()
        # Create a log entry for the proposal
        instance.conservation_status.log_user_action(
            ConservationStatusUserAction.COMMENT_REFERRAL.format(
                instance.id,
                instance.conservation_status.conservation_status_number,
                f"{instance.referral_as_email_user.get_full_name()}({instance.referral_as_email_user.email})",
            ),
            request,
        )
        return redirect(reverse("internal"))


class ConservationStatusAmendmentRequestViewSet(
    viewsets.GenericViewSet, mixins.RetrieveModelMixin
):
    queryset = ConservationStatusAmendmentRequest.objects.none()
    serializer_class = ConservationStatusAmendmentRequestSerializer

    def get_queryset(self):
        if is_internal(self.request):  # user.is_authenticated():
            qs = ConservationStatusAmendmentRequest.objects.all().order_by("id")
            return qs
        return ConservationStatusAmendmentRequest.objects.none()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=json.loads(request.data.get("data")))
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        instance.add_documents(request)
        instance.generate_amendment(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    @renderer_classes((JSONRenderer,))
    def delete_document(self, request, *args, **kwargs):
        instance = self.get_object()
        ConservationStatusAmendmentRequestDocument.objects.get(
            id=request.data.get("id")
        ).delete()
        return Response(
            [
                dict(id=i.id, name=i.name, _file=i._file.url)
                for i in instance.cs_amendment_request_documents.all()
            ]
        )


class ConservationStatusDocumentViewSet(
    viewsets.GenericViewSet, mixins.RetrieveModelMixin
):
    queryset = ConservationStatusDocument.objects.none()
    serializer_class = ConservationStatusDocumentSerializer

    def get_queryset(self):
        if is_internal(self.request):
            return ConservationStatusDocument.objects.all().order_by("id")
        if is_external_contributor(self.request):
            return ConservationStatusDocument.objects.filter(
                conservation_status__submitter=self.request.user.id,
                visible=True,
                can_submitter_access=True,
            )
        return ConservationStatusDocument.objects.none()

    def get_serializer_class(self):
        return super().get_serializer_class()

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.visible = False
        instance.save(version_user=request.user)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.visible = True
        instance.save(version_user=request.user)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        data = json.loads(request.data.get("data"))
        serializer = SaveConservationStatusDocumentSerializer(instance, data=data)
        if is_internal(self.request):
            serializer = InternalSaveConservationStatusDocumentSerializer(
                instance, data=data
            )

        serializer.is_valid(raise_exception=True)
        serializer.save(no_revision=True)
        instance.add_documents(request, version_user=request.user)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        data = json.loads(request.data.get("data"))
        serializer = SaveConservationStatusDocumentSerializer(data=data)
        if is_internal(self.request):
            serializer = InternalSaveConservationStatusDocumentSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save(no_revision=True)
        instance.add_documents(request, version_user=request.user)
        return Response(serializer.data)


class AmendmentRequestReasonChoicesView(views.APIView):

    renderer_classes = [
        JSONRenderer,
    ]

    def get(self, request, format=None):
        choices_list = []
        choices = ProposalAmendmentReason.objects.all()
        if choices:
            for c in choices:
                choices_list.append({"key": c.id, "value": c.reason})
        return Response(choices_list)
